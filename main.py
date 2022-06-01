import requests
from re import search
import openpyxl
from datetime import datetime
from bs4 import BeautifulSoup
page = requests.get("https://www.mirea.ru/schedule/")
soup = BeautifulSoup(page.text, "html.parser")
result = soup.find("div", {"class":"rasspisanie"}).\
    find(string = "Институт информационных технологий").\
    find_parent("div").\
    find_parent("div").\
    find_all('a', class_='uk-link-toggle')

links = []

for x in result:
    links.append(x['href'])

i = 1
"""
for link in links:
    if "/ИИТ_1" in link: # среди всех ссылок найти нужную
        print (link)
        f = open('kurs_1.xlsx', "wb") # открываем файл для записи, в режиме wb
        resp = requests.get(link) # запрос по ссылке
        f.write(resp.content)
        i+=1
"""
book = openpyxl.load_workbook("kurs_1.xlsx")  # открытие файла
sheet = book.active  # активный лист

user_message = "ИКБО-09-21"
wek=0
wekday=1
list_par = []

def shed_day(list_par, x, y, wek):
    list_par = []
    for col in sheet.iter_cols(min_row=2, max_col=sheet.max_column, max_row=2):
        for cell in col:
            if search(r'([А-Я]{4}-\d{2}-\d{2})', str(cell.value)):
                if cell.value==user_message:
                    colm=cell.column
                    for row in range(x+wek,y,2):
                        list_par.append(sheet.cell(row=row,column=colm).value)
    return list_par




if (wekday==0):
    shed_day(list_par, 4, 15, wek)
elif (wekday==1):
    shed_day(list_par, 16, 27, wek)
elif (wekday==2):
    shed_day(list_par, 28, 39, wek)
elif (wekday==3):
    shed_day(list_par, 40, 51, wek)
elif (wekday==4):
    shed_day(list_par, 52, 63, wek)
elif (wekday==5):
    shed_day(list_par, 64, 75, wek)
else:
    list_par = []


    def _get_number_week(day_today: datetime) -> int:
        """Получение номера недели.

        Parameters
        ----------
        day_today: datetime
            сегодняшняя дата
        Return
        ----------
        number: int
            номер недели
        """
        first_week = Config.get_weeks_info()["start_week"].isocalendar()[1]
        current_week = day_today.isocalendar()[1]
        number = current_week - first_week + 1
        return number

print(self._get_number_week(datetime.today()))